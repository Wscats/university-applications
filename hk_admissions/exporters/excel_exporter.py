"""
Excel exporter for Hong Kong university admission data.

Exports programme data to .xlsx format using openpyxl.
"""

import logging
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from hk_admissions.models import ProgramInfo

logger = logging.getLogger(__name__)

# Column definitions: (header, field_name, width)
COLUMNS = [
    ("University", "university_name", 35),
    ("University (CN)", "university_name_cn", 15),
    ("Abbreviation", "university_abbreviation", 12),
    ("Faculty / School", "faculty", 30),
    ("Programme Name", "program_name", 45),
    ("Programme Name (CN)", "program_name_cn", 20),
    ("Degree Type", "degree_type", 12),
    ("Tuition Fee", "tuition_fee", 20),
    ("Currency", "tuition_currency", 10),
    ("Tuition Remarks", "tuition_remarks", 25),
    ("Application Open Date", "application_open_date", 20),
    ("Application Deadline", "application_deadline", 20),
    ("Deadline Remarks", "application_deadline_remarks", 30),
    ("English Requirement", "english_requirement", 40),
    ("English Details", "english_requirement_details", 30),
    ("Duration", "duration", 20),
    ("Mode", "mode", 18),
    ("Official Programme URL", "program_url", 50),
    ("Data Source", "data_source", 50),
    ("Last Updated", "last_updated", 15),
    ("Remarks", "remarks", 30),
]


class ExcelExporter:
    """Export admission data to Excel (.xlsx) format."""

    def export(
        self,
        programs: List[ProgramInfo],
        output_dir: str,
        filename_prefix: str = "hk_master_admissions",
    ) -> str:
        """
        Export programs to an Excel file.

        Args:
            programs: List of ProgramInfo objects.
            output_dir: Directory to save the file.
            filename_prefix: Filename prefix.

        Returns:
            Path to the created Excel file.
        """
        wb = Workbook()

        # ---- Sheet 1: All Programs ----
        ws_all = wb.active
        ws_all.title = "All Programs"
        self._write_sheet(ws_all, programs)

        # ---- Sheet 2: Summary by University ----
        ws_summary = wb.create_sheet("Summary")
        self._write_summary(ws_summary, programs)

        # ---- Per-university sheets ----
        universities = {}
        for p in programs:
            key = p.university_abbreviation or p.university_name or "Other"
            if key not in universities:
                universities[key] = []
            universities[key].append(p)

        for uni_key, uni_programs in sorted(universities.items()):
            # Sheet name max 31 chars
            sheet_name = uni_key[:31]
            ws = wb.create_sheet(sheet_name)
            self._write_sheet(ws, uni_programs)

        # Save
        filepath = str(Path(output_dir) / f"{filename_prefix}.xlsx")
        wb.save(filepath)
        logger.info(f"Excel file saved: {filepath}")
        return filepath

    def _write_sheet(self, ws, programs: List[ProgramInfo]):
        """Write programmes data to a worksheet."""
        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Write data
        data_alignment = Alignment(vertical="top", wrap_text=True)
        for row_idx, program in enumerate(programs, 2):
            for col_idx, (_, field, _) in enumerate(COLUMNS, 1):
                value = getattr(program, field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border

                # Make URLs clickable
                if field == "program_url" and value:
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-filter
        if programs:
            ws.auto_filter.ref = ws.dimensions

    def _write_summary(self, ws, programs: List[ProgramInfo]):
        """Write summary statistics sheet."""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Summary headers
        headers = ["University", "Chinese Name", "Abbreviation", "Program Count", "Admissions URL"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 55

        # Aggregate
        uni_data = {}
        for p in programs:
            key = p.university_name
            if key not in uni_data:
                uni_data[key] = {
                    "name_cn": p.university_name_cn,
                    "abbr": p.university_abbreviation,
                    "count": 0,
                    "url": p.data_source or p.program_url,
                }
            uni_data[key]["count"] += 1

        row = 2
        for name, data in sorted(uni_data.items()):
            ws.cell(row=row, column=1, value=name).border = thin_border
            ws.cell(row=row, column=2, value=data["name_cn"]).border = thin_border
            ws.cell(row=row, column=3, value=data["abbr"]).border = thin_border
            ws.cell(row=row, column=4, value=data["count"]).border = thin_border
            url_cell = ws.cell(row=row, column=5, value=data["url"])
            url_cell.border = thin_border
            if data["url"]:
                url_cell.hyperlink = data["url"]
                url_cell.font = Font(color="0563C1", underline="single")
            row += 1

        # Total row
        total_row = row
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=len(programs)).font = Font(bold=True)

        ws.freeze_panes = "A2"
